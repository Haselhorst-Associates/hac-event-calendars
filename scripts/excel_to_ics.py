"""
excel_to_ics.py — Excel Event Calendar → ICS Converter

Reads event data from .xlsx files and generates RFC 5545 compliant ICS files.
Supports EN and DE column headers, multi-day date ranges, deterministic UIDs.

Usage:
    # CLI (for GitHub Actions / terminal)
    python excel_to_ics.py --config calendars.yaml --base-dir .

    # Python import
    from excel_to_ics import process_all_calendars
    results = process_all_calendars("calendars.yaml", base_dir=".")
"""

import argparse
import json
import logging
import re
import uuid
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl
import yaml

logger = logging.getLogger(__name__)

# ── Column mappings ──────────────────────────────────────────────────────────

COLUMN_MAP = {
    "en": {
        "date": "Date",
        "event": "Event",
        "organiser": "Organiser",
        "location": "Location",
        "country": "Country",
        "cost": "Cost",
        "link": "Link",
        "status": "Status",
        "category": "Category",
    },
    "de": {
        "date": "Datum",
        "event": "Veranstaltung",
        "organiser": "Veranstalter",
        "location": "Ort",
        "country": "Land",
        "cost": "Kosten",
        "link": "Link",
        "status": "Status",
        "category": "Kategorie",
    },
}

# Description labels per language
LABELS = {
    "en": {"cost": "Cost", "link_label": "Registration & Info"},
    "de": {"cost": "Kosten", "link_label": "Anmeldung & Infos"},
}

# UUID namespace for deterministic UIDs
HAC_NAMESPACE = uuid.UUID("a3e7b8c1-4d5f-6e7a-8b9c-0d1e2f3a4b5c")

# ── Date parsing ─────────────────────────────────────────────────────────────

# Matches: "03.–05.02.2026", "31.08.–01.09.2026", "02.02.2026"
# Handles hyphen (-), en-dash (–), em-dash (—)
DATE_RANGE_PATTERN = re.compile(
    r"^(\d{1,2})\.\s*[–—-]\s*(\d{1,2})\.(\d{2})\.(\d{4})$"
)
DATE_RANGE_CROSS_MONTH_PATTERN = re.compile(
    r"^(\d{1,2})\.(\d{2})\.\s*[–—-]\s*(\d{1,2})\.(\d{2})\.(\d{4})$"
)
DATE_SINGLE_PATTERN = re.compile(
    r"^(\d{1,2})\.(\d{2})\.(\d{4})$"
)


def parse_date_range(date_str: str) -> tuple[date, date] | None:
    """Parse a date string into (start_date, end_date_exclusive).

    DTEND is always day+1 for all-day events per RFC 5545.
    Returns None if the date cannot be parsed (e.g. "September 2026 (tbd)").
    """
    if not date_str or not isinstance(date_str, str):
        return None

    s = date_str.strip()

    # Single day: "02.02.2026"
    m = DATE_SINGLE_PATTERN.match(s)
    if m:
        day, month, year = int(m.group(1)), int(m.group(2)), int(m.group(3))
        start = date(year, month, day)
        return start, start + timedelta(days=1)

    # Range within same month: "03.–05.02.2026"
    m = DATE_RANGE_PATTERN.match(s)
    if m:
        day_start = int(m.group(1))
        day_end = int(m.group(2))
        month = int(m.group(3))
        year = int(m.group(4))
        start = date(year, month, day_start)
        end = date(year, month, day_end) + timedelta(days=1)
        return start, end

    # Range across months: "31.08.–01.09.2026"
    m = DATE_RANGE_CROSS_MONTH_PATTERN.match(s)
    if m:
        day_start = int(m.group(1))
        month_start = int(m.group(2))
        day_end = int(m.group(3))
        month_end = int(m.group(4))
        year = int(m.group(5))
        start = date(year, month_start, day_start)
        end = date(year, month_end, day_end) + timedelta(days=1)
        return start, end

    return None


# ── ICS text helpers ─────────────────────────────────────────────────────────

def _ics_escape(text: str) -> str:
    """Escape text for ICS content lines (RFC 5545 §3.3.11)."""
    return text.replace("\\", "\\\\").replace(";", "\\;").replace(",", "\\,").replace("\n", "\\n")


def _fold_line(line: str) -> str:
    """Fold long lines per RFC 5545 (max 75 octets per line).

    Continuation lines start with a single space.
    """
    encoded = line.encode("utf-8")
    if len(encoded) <= 75:
        return line

    parts = []
    while len(encoded) > 75:
        # Find a safe cut point (don't break multi-byte UTF-8)
        cut = 75 if not parts else 74  # first line: 75, continuation: 74 (space prefix)
        while cut > 0 and (encoded[cut] & 0xC0) == 0x80:
            cut -= 1
        if not parts:
            parts.append(encoded[:cut].decode("utf-8"))
        else:
            parts.append(" " + encoded[:cut].decode("utf-8"))
        encoded = encoded[cut:]

    if encoded:
        if parts:
            parts.append(" " + encoded.decode("utf-8"))
        else:
            parts.append(encoded.decode("utf-8"))

    return "\r\n".join(parts)


def _format_date(d: date) -> str:
    """Format date as ICS VALUE=DATE string."""
    return d.strftime("%Y%m%d")


# ── Excel reading ────────────────────────────────────────────────────────────

def read_excel_events(
    excel_path: str | Path,
    header_row: int,
    language: str,
) -> list[dict]:
    """Read events from an Excel file.

    Args:
        excel_path: Path to the .xlsx file.
        header_row: 0-indexed row number containing column headers.
        language: "en" or "de" — determines column name mapping.

    Returns:
        List of event dicts with normalized keys.
    """
    wb = openpyxl.load_workbook(str(excel_path), data_only=True)
    ws = wb.active

    # Build reverse column map: Excel header → internal key
    col_map = COLUMN_MAP[language]
    reverse_map = {v: k for k, v in col_map.items()}

    # Read header row (0-indexed → openpyxl 1-indexed)
    header_row_1idx = header_row + 1
    headers = []
    for cell in ws[header_row_1idx]:
        val = cell.value
        if val and isinstance(val, str):
            headers.append((cell.column - 1, reverse_map.get(val.strip(), None)))
        else:
            headers.append((cell.column - 1, None))

    # Map column index → internal key
    col_idx_to_key = {idx: key for idx, key in headers if key is not None}

    events = []
    for row in ws.iter_rows(min_row=header_row_1idx + 1, max_row=ws.max_row):
        record = {}
        for cell in row:
            key = col_idx_to_key.get(cell.column - 1)
            if key:
                record[key] = cell.value

        # Skip empty rows
        if not record.get("event") or not record.get("date"):
            continue

        # Normalize date to string
        date_val = record["date"]
        if isinstance(date_val, (datetime, date)):
            # Excel stored it as a date object → convert to DD.MM.YYYY
            record["date"] = date_val.strftime("%d.%m.%Y")
        elif isinstance(date_val, str):
            record["date"] = date_val.strip()
        else:
            record["date"] = str(date_val)

        # Strip whitespace from string fields
        for k, v in record.items():
            if isinstance(v, str):
                record[k] = v.strip()

        events.append(record)

    wb.close()
    return events


# ── ICS generation ───────────────────────────────────────────────────────────

def _make_uid(summary: str, dtstart: date) -> str:
    """Generate a deterministic UUID5 from event name + start date."""
    seed = f"{summary}|{_format_date(dtstart)}"
    return str(uuid.uuid5(HAC_NAMESPACE, seed))


def _build_description(event: dict, language: str) -> tuple[str, str]:
    """Build plain text and HTML description from event data.

    Returns:
        (plain_text_description, html_description)
    """
    labels = LABELS[language]
    cost = event.get("cost", "")
    link = event.get("link", "")

    # Plain text (ICS DESCRIPTION)
    parts = []
    if cost:
        parts.append(f"{labels['cost']}: {cost}")
    if link:
        parts.append(f"{labels['link_label']}: {link}")
    plain = "\\n".join(parts)

    # HTML (X-ALT-DESC)
    html_parts = []
    if cost:
        html_parts.append(f"<p>{labels['cost']}: {_html_escape(cost)}</p>")
    if link:
        html_parts.append(
            f'<p><a href="{_html_escape(link)}">{labels["link_label"]}</a></p>'
        )
    html = f"<!DOCTYPE HTML><HTML><BODY>{''.join(html_parts)}</BODY></HTML>"

    return plain, html


def _html_escape(text: str) -> str:
    """Minimal HTML escaping for ICS X-ALT-DESC."""
    return text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")


def _build_summary(event: dict, summary_prefix: str) -> str:
    """Build SUMMARY field, optionally with country prefix."""
    name = event.get("event", "")
    country = event.get("country", "")

    if summary_prefix == "country" and country:
        return f"[{country}] {name}"
    return name


def _build_location(event: dict) -> str:
    """Build LOCATION field from location + country."""
    location = event.get("location", "")
    return location if location else ""


def event_to_vevent(event: dict, cal_config: dict) -> str | None:
    """Convert a single event dict to a VEVENT block.

    Returns None if the event date cannot be parsed.
    """
    language = cal_config.get("language", "en")
    summary_prefix = cal_config.get("summary_prefix", "none")

    # Parse date
    parsed = parse_date_range(event.get("date", ""))
    if parsed is None:
        logger.warning(
            "Skipping event with unparseable date: %s — %s",
            event.get("date", "?"),
            event.get("event", "?"),
        )
        return None

    dtstart, dtend = parsed

    summary = _build_summary(event, summary_prefix)
    location = _build_location(event)
    description_plain, description_html = _build_description(event, language)
    uid = _make_uid(summary, dtstart)
    url = event.get("link", "")
    category = event.get("category", "")

    lines = [
        "BEGIN:VEVENT",
        f"UID:{uid}",
        f"DTSTART;VALUE=DATE:{_format_date(dtstart)}",
        f"DTEND;VALUE=DATE:{_format_date(dtend)}",
        f"SUMMARY:{_ics_escape(summary)}",
    ]

    if location:
        lines.append(f"LOCATION:{_ics_escape(location)}")

    if description_plain:
        lines.append(f"DESCRIPTION:{description_plain}")

    if url:
        lines.append(f"URL:{url}")

    if description_html:
        lines.append(f"X-ALT-DESC;FMTTYPE=text/html:{description_html}")

    if category:
        lines.append(f"CATEGORIES:{_ics_escape(category)}")

    lines.append("TRANSP:TRANSPARENT")
    lines.append("END:VEVENT")

    return "\r\n".join(_fold_line(line) for line in lines)


def generate_ics(events: list[dict], cal_config: dict) -> str:
    """Generate a complete ICS calendar string from a list of events."""
    calendar_name = cal_config.get("calendar_name", "Events")
    prodid = cal_config.get("prodid", "-//HAC//Events//EN")

    header_lines = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        f"PRODID:{prodid}",
        "CALSCALE:GREGORIAN",
        "METHOD:PUBLISH",
        f"X-WR-CALNAME:{calendar_name}",
        "X-WR-TIMEZONE:Europe/Berlin",
    ]

    header = "\r\n".join(header_lines)
    footer = "END:VCALENDAR"

    # Sort events chronologically by start date
    def _sort_key(event):
        parsed = parse_date_range(event.get("date", ""))
        return parsed[0] if parsed else date.max

    events_sorted = sorted(events, key=_sort_key)

    vevent_blocks = []
    skipped = 0
    for event in events_sorted:
        block = event_to_vevent(event, cal_config)
        if block:
            vevent_blocks.append(block)
        else:
            skipped += 1

    body = "\r\n".join(vevent_blocks)
    ics_content = f"{header}\r\n{body}\r\n{footer}\r\n"

    logger.info(
        "Generated %d events for '%s' (%d skipped)",
        len(vevent_blocks),
        calendar_name,
        skipped,
    )

    return ics_content


# ── Pipeline ─────────────────────────────────────────────────────────────────

def load_config(config_path: str | Path) -> dict:
    """Load YAML configuration file."""
    with open(config_path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def process_calendar(cal_config: dict, base_dir: Path) -> dict:
    """Process a single calendar: read Excel → generate ICS → write file.

    Returns:
        Stats dict with event counts.
    """
    excel_path = base_dir / cal_config["excel_file"]
    output_path = base_dir / cal_config["output_file"]
    header_row = cal_config.get("header_row", 3)
    language = cal_config.get("language", "en")

    logger.info("Reading %s (language=%s, header_row=%d)", excel_path.name, language, header_row)

    # Read events
    events = read_excel_events(excel_path, header_row, language)
    logger.info("Found %d events in %s", len(events), excel_path.name)

    # Filter by status if configured
    skip_status = cal_config.get("skip_status", [])
    if skip_status:
        before = len(events)
        events = [e for e in events if e.get("status", "") not in skip_status]
        logger.info("Filtered %d events by status (skip: %s)", before - len(events), skip_status)

    # Generate ICS
    ics_content = generate_ics(events, cal_config)

    # Write output
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8", newline="") as f:
        f.write(ics_content)

    logger.info("Wrote %s", output_path)

    return {
        "excel_file": str(excel_path.name),
        "output_file": str(output_path),
        "total_events": len(events),
        "calendar_name": cal_config.get("calendar_name", ""),
    }


def process_all_calendars(config_path: str | Path, base_dir: str | Path = ".") -> list[dict]:
    """Process all calendars defined in the YAML config.

    Args:
        config_path: Path to calendars.yaml.
        base_dir: Base directory for resolving relative paths in config.

    Returns:
        List of stats dicts, one per calendar.
    """
    config = load_config(config_path)
    base = Path(base_dir)
    results = []

    for cal_config in config.get("calendars", []):
        try:
            stats = process_calendar(cal_config, base)
            results.append(stats)
        except Exception:
            logger.exception("Failed to process calendar: %s", cal_config.get("excel_file", "?"))
            results.append({
                "excel_file": cal_config.get("excel_file", "?"),
                "error": True,
            })

    return results


# ── CLI entry point ──────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Convert Excel event calendars to ICS files."
    )
    parser.add_argument(
        "--config",
        required=True,
        help="Path to calendars.yaml configuration file.",
    )
    parser.add_argument(
        "--base-dir",
        default=".",
        help="Base directory for resolving relative paths (default: current dir).",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Enable verbose logging.",
    )
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(levelname)s: %(message)s",
    )

    results = process_all_calendars(args.config, args.base_dir)
    print(json.dumps(results, ensure_ascii=False, indent=2))
